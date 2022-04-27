import MySQLdb
from openpyxl import load_workbook

db = MySQLdb.connect("localhost", "root", "admin", "reactflask")
print(db, "_______________________")
cursor = db.cursor()
global resultsExport
resultsExport = []
resultsExport2 = []
resultsExport3 = []


def get_hello_world():
    del resultsExport3[:]
    sql = "SELECT * FROM users"
    try:
        cursor.execute(sql)
        results = cursor.fetchall()
        for row in results:
            item = {
                "Id": row[0],
                "username": row[1],
                "email": row[2],
                
            }
            resultsExport3.append(item)
        return resultsExport3
    except MySQLdb.Error as e:
        try:
            msg = "MySQL Error [%d]: %s" % (e.args[0], e.args[1])
            return msg
        except IndexError:
            msg =  "MySQL Error: %s" % str(e)
            return msg
        finally:
            cursor.close()
            db.close()


def get_ErrorLog():
    del resultsExport2[:]
    sql = "SELECT * FROM errorlog"
    try:
        cursor.execute(sql)
        results = cursor.fetchall()
        for row in results:
            item = {
                "iderrorlog": row[0],
                "ScriptNameUser": row[1],
                "CreateTime": row[2]
            }
            resultsExport2.append(item)
        return resultsExport2
    except MySQLdb.Error as e:
        try:
            msg = "MySQL Error [%d]: %s" % (e.args[0], e.args[1])
            return msg
        except IndexError:
            msg =  "MySQL Error: %s" % str(e)
            return msg
        finally:
            cursor.close()
            db.close()

path2 = "C:\\Users\\Oumaima\\Downloads\\ErrorLog.xlsx"
wb2 = load_workbook(path2)
sheet2 = wb2.active
m_row2 = sheet2.max_row

def create_ErrorLog():
    val = []
    for j in range(2,m_row2 + 1 ):

        iderrorlog = sheet2.cell(row=j, column=1).value
        ScriptNameUser = sheet2.cell(row=j, column=6).value
        CreateTime = sheet2.cell(row=j, column=15).value
        values = (iderrorlog ,  ScriptNameUser ,CreateTime)
        val.append(values)
    sql = "Insert into errorlog (iderrorlog,  ScriptNameUser, CreateTime) values (%s, %s, %s)"
    try:
        cursor.execute("DELETE FROM errorlog")
        cursor.executemany(sql,val)
        db.commit()
        return val
    except MySQLdb.Error as e:
        try:
            msg="MySQL Error [%d]: %s" % (e.args[0], e.args[1])
            return msg
        except IndexError:
            db.rollback()
            msg= "MySQL Error: %s" % str(e)
            return msg

def get_rew_log():
    del resultsExport[:]
    sql = "SELECT * FROM rew_log"
    try:
        cursor.execute(sql)
        results = cursor.fetchall()
        for row in results:
            item = {
                "idREW_Log": row[0],
                "KSK": row[1],
                "operator_comment": row[2],
                "CreateTime": row[3]
            }
            resultsExport.append(item)
        return resultsExport
    except MySQLdb.Error as e:
        try:
            msg = "MySQL Error [%d]: %s" % (e.args[0], e.args[1])
            return msg
        except IndexError:
            msg =  "MySQL Error: %s" % str(e)
            return msg
        finally:
            cursor.close()
            db.close()



path = "C:\\Users\\Oumaima\\Downloads\\REW_Log.xlsx"
wb = load_workbook(path)
sheet = wb.active
m_row = sheet.max_row

def create_rew_log():
    val = []
    for j in range(2,m_row + 1 ):
        idREW_Log = sheet.cell(row=j, column=1).value
        ksk = sheet.cell(row=j, column=2).value
        OperatorComment = sheet.cell(row=j, column=15).value
        CreateTime = sheet.cell(row=j, column=19).value
        values = (idREW_Log , ksk, OperatorComment,CreateTime)
        val.append(values)
    sql = "Insert into rew_log (idREW_Log, KSK, operator_comment, CreateTime) values (%s, %s, %s, %s)"
    try:
        cursor.execute("DELETE FROM rew_log")
        cursor.executemany(sql,val)
        db.commit()
        return val
    except MySQLdb.Error as e:
        try:
            msg="MySQL Error [%d]: %s" % (e.args[0], e.args[1])
            return msg
        except IndexError:
            db.rollback()
            msg= "MySQL Error: %s" % str(e)
            return msg
    

